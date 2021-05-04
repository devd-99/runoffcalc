import matplotlib.pyplot as plt
import numpy as np
import openpyxl

wbkName = 'GRUN_Smaktin.xlsx'
wbk = openpyxl.load_workbook(wbkName)

timeArray=[]
runoffArray = []
colStart = 5
qiArray = []

for wks in wbk.worksheets:
    while(wks.cell(row=39, column=colStart).value is not None):
        
        runoffArray.append(wks.cell(row=36, column=colStart).value)
        timeArray.append(wks.cell(row=39, column=colStart).value)
        qiArray.append(wks.cell(row=41, column=colStart).value)
        colStart+=1

# graph 1
fig, ax = plt.subplots()  # Create a figure containing a single axes.
ax.plot(runoffArray, timeArray)  # Plot some data on the axes.
ax.xaxis.set_major_locator(plt.MaxNLocator(10))
ax.yaxis.set_major_locator(plt.MaxNLocator(10))
plt.xlabel("Time")
plt.ylabel("Runoff (m3/s)")
plt.savefig('plot1.png', dpi=300)

# graph 2
fig2, ax2 = plt.subplots()  # Create a figure containing a single axes.
ax2.plot(qiArray, runoffArray)  # Plot some data on the axes.
ax2.xaxis.set_major_locator(plt.MaxNLocator(10))
ax2.yaxis.set_major_locator(plt.MaxNLocator(10))
ax2.set_yscale('log')
plt.xlabel("qi")
plt.ylabel("Runoff (m3/s)")
plt.savefig('plot2.png', dpi=300)

wbk.close()