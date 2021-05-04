import openpyxl

wbkName = 'GRUN_Smaktin.xlsx'
wbk = openpyxl.load_workbook(wbkName)

colStart = 5


for wks in wbk.worksheets:

    while(wks.cell(row=37, column = colStart).value is not None):

        runoffArray = []
        seq=[]
        index=[]
        qiVals= []

        i=1
        for col in range(colStart, colStart+12):
            wks.cell(row=36, column=col).value = str(i)+"/"+str(wks.cell(row=37, column = colStart).value)
            runoffArray.append(wks.cell(row=39, column=col).value)
            i+=1

        seq = sorted(runoffArray, reverse=True)
        index= [(seq.index(v)+1) for v in runoffArray]

        qiVals = [(v*100)/12 for v in index]

        print(colStart)
        i=0
        for col in range(colStart, colStart+12):
            
            wks.cell(row=40, column=col).value = index[i]
            wks.cell(row=41, column=col).value = qiVals[i]
            i+=1

        colStart=colStart+12
wbk.save(wbkName)

wbk.close()